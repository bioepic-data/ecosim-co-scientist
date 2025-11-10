#!/usr/bin/env python
"""Quick script to explore the benchmark data."""

import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('hackathon-case_study-experimental_warming_nitrogen/experimental_warming_nitrogen-benchmark_data.xlsx')

print('Available sheets:', wb.sheetnames)
print()

# Examine each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    print(f"Dimensions: {ws.dimensions}")
    print("\nFirst 6 rows:")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 5:
            break
        print(row)
